from PySide6.QtCore import QAbstractTableModel, Qt
from PySide6.QtGui import QFont
from openpyxl import Workbook
from openpyxl.styles import Font as ExcelFont
from openpyxl.utils import get_column_letter

from app.core.cliente_woocommerce import ClienteWooCommerce
from app.menu.menu_view import MenuPrincipalView


HEADERS = [
    "FECHA", "CLIENTE", "SUBTOTAL", "ENVÍO", "IVA", "DESCUENTO", "TOTAL",
    "UTILIDAD", "ESTADO", "NOTAS", "PEDIDO", "IDENTIFICACIÓN",
    "CORREO", "TELÉFONO", "DIRECCIÓN", "CIUDAD", "CAJERO"
]

ESTADOS_ES = {
    "pending": "Pendiente",
    "processing": "Procesando",
    "on-hold": "En espera",
    "completed": "Completado",
    "cancelled": "Cancelado",
    "refunded": "Reembolsado",
    "failed": "Fallido",
}


class ModeloTablaReporteVentas(QAbstractTableModel):
    def __init__(self, datos):
        super().__init__()
        self._datos = datos

    def rowCount(self, parent=None):
        return len(self._datos)

    def columnCount(self, parent=None):
        return len(HEADERS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == Qt.DisplayRole:
            return self._datos[index.row()][index.column()]

        if role == Qt.TextAlignmentRole:
            return Qt.AlignCenter

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal:
            if role == Qt.DisplayRole:
                return HEADERS[section]
            if role == Qt.FontRole:
                f = QFont()
                f.setBold(True)
                return f
            if role == Qt.TextAlignmentRole:
                return Qt.AlignCenter
        return None


class ControladorReporteVentas:
    def __init__(self):
        self.cliente = ClienteWooCommerce()
        self._datos = []

    def _obtener_cajero(self, pedido):
        for meta in pedido.get("meta_data", []):
            key = meta.get("key", "").lower()
            if key in ("cajero", "cashier", "vendedor", "pos_user"):
                return str(meta.get("value", "") or "")
        return ""

    def generar_reporte(self, desde, hasta, callback_progreso=None):
        pedidos = self.cliente.obtener_pedidos(
            desde=desde.isoformat(),
            hasta=hasta.isoformat(),
            per_page=100
        )

        self._datos.clear()
        total = len(pedidos) or 1

        for i, pedido in enumerate(pedidos, start=1):
            self._datos.append(self._procesar_pedido(pedido))

            if callback_progreso:
                callback_progreso(
                    int((i / total) * 100),
                    f"Procesando orden {i} de {total}"
                )

        return ModeloTablaReporteVentas(self._datos)

    def _procesar_pedido(self, pedido):
        billing = pedido.get("billing", {})
        shipping = pedido.get("shipping", {})

        subtotal = float(pedido.get("subtotal", 0))
        total = float(pedido.get("total", 0))
        iva = float(pedido.get("total_tax", 0))
        envio = float(pedido.get("shipping_total", 0))
        descuento = float(pedido.get("discount_total", 0))

        utilidad = round(total - descuento, 2)

        estado = ESTADOS_ES.get(pedido.get("status"), pedido.get("status", ""))

        direccion = " ".join(filter(None, [
            shipping.get("address_1", ""),
            shipping.get("address_2", "")
        ]))

        return [
            pedido.get("date_created", "")[:10],
            f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip(),
            round(subtotal, 2),
            round(envio, 2),
            round(iva, 2),
            round(descuento, 2),
            round(total, 2),
            utilidad,
            estado,
            pedido.get("customer_note", ""),
            pedido.get("number", ""),
            billing.get("company", ""),
            billing.get("email", ""),
            billing.get("phone", ""),
            direccion,
            shipping.get("city", ""),
            self._obtener_cajero(pedido),
        ]

    def exportar_excel(self, ruta):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Ventas"

        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelFont(bold=True)

        for row_idx, fila in enumerate(self._datos, start=2):
            for col_idx, valor in enumerate(fila, start=1):
                ws.cell(row=row_idx, column=col_idx, value=valor)

        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(ruta)
