# app/actualizar_productos/controlador_actualizar_productos.py
import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt
from PySide6.QtGui import QFont

from app.core.cliente_woocommerce import ClienteWooCommerce


HEADERS_UI = [
    "SKU", "NOMBRE", "CATEGORÍA",
    "STOCK ACTUAL", "STOCK NUEVO",
    "PRECIO ACTUAL", "PRECIO COMPRA",
    "PRECIO VENTA ACTUAL", "PRECIO VENTA NUEVO",
    "ESTADO"
]

COL_SKU = 0
COL_NOMBRE = 1
COL_CATEGORIA = 2
COL_STOCK_ACTUAL = 3
COL_STOCK_NUEVO = 4
COL_PRECIO_ACTUAL = 5
COL_PRECIO_COMPRA = 6
COL_PRECIO_VENTA_ACTUAL = 7
COL_PRECIO_VENTA_NUEVO = 8
COL_ESTADO = 9


# ----------------------------
# Utilidades
# ----------------------------
def _norm_header(h: Any) -> str:
    """Normaliza encabezados: str, strip, upper, espacios simples."""
    if h is None:
        return ""
    return " ".join(str(h).strip().upper().split())


def _leer_num_opcional(valor: Any) -> Optional[float]:
    """Devuelve float o None si viene vacío. Acepta coma decimal."""
    if valor is None:
        return None
    txt = str(valor).strip()
    if txt == "":
        return None
    txt = txt.replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return None


def _mapear_columna(headers: List[str], aliases: List[str], requerido: bool, nombre_para_error: str) -> Optional[int]:
    """
    Devuelve índice de la primera coincidencia entre aliases.
    Si requerido=True, lanza error si no existe.
    """
    aliases_norm = [_norm_header(a) for a in aliases]
    for a in aliases_norm:
        if a in headers:
            return headers.index(a)
    if requerido:
        raise ValueError(
            f"Falta la columna requerida '{nombre_para_error}'. "
            f"Se aceptan: {', '.join(aliases)}"
        )
    return None


def _validar_negrita_excel(ws, idxs_requeridos: List[int], nombres_requeridos: List[str]):
    """Valida que los encabezados requeridos (fila 1) estén en negrita."""
    for idx, nombre in zip(idxs_requeridos, nombres_requeridos):
        celda = ws.cell(row=1, column=idx + 1)  # openpyxl usa 1-based
        bold = bool(getattr(celda.font, "bold", False))
        if not bold:
            raise ValueError(f"En Excel, el encabezado '{nombre}' debe estar en NEGRITA.")


def _to_decimal(x: Any) -> Decimal:
    if x is None:
        return Decimal("0")
    s = str(x).strip()
    if not s:
        return Decimal("0")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _fmt_money(x: Any) -> str:
    """
    Máximo 2 decimales (redondeo financiero).
    - 12.00 -> 12
    - 6.50 -> 6.5
    - 1.7391 -> 1.74
    """
    d = _to_decimal(x).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = format(d, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def _get_meta(meta_data: List[dict], key: str) -> Optional[str]:
    for m in meta_data or []:
        if m.get("key") == key:
            v = m.get("value")
            return None if v is None else str(v)
    return None


def _get_purchase_cost(product: dict) -> Decimal:
    """
    Prioridad multi-tienda:
      1) _purchase_price (ATUM)
      2) _wc_cog_cost (Cost of Goods)
      3) purchase_price (legacy/custom)
    """
    meta = product.get("meta_data", []) or []
    for k in ("_purchase_price", "_wc_cog_cost", "purchase_price"):
        v = _get_meta(meta, k)
        if v is not None and _to_decimal(v) > 0:
            return _to_decimal(v)
    return Decimal("0")


def _to_float_any(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


# ----------------------------
# Modelo de datos
# ----------------------------
@dataclass
class RegistroProducto:
    sku: str
    nombre: str
    categoria: str
    stock_actual: Any
    stock_nuevo: Optional[int]
    precio_actual: Any
    precio_compra: Any
    precio_venta_actual: Any
    precio_venta_nuevo: Optional[float]
    estado: str = ""

    _id: Optional[int] = None
    _parent_id: Optional[int] = None
    _tipo: str = "simple"

    def como_fila(self) -> List[Any]:
        return [
            self.sku,
            self.nombre,
            self.categoria,
            self.stock_actual,
            "" if self.stock_nuevo is None else self.stock_nuevo,
            self.precio_actual,
            self.precio_compra,
            self.precio_venta_actual,
            "" if self.precio_venta_nuevo is None else self.precio_venta_nuevo,
            self.estado,
        ]

    def tiene_cambios(self) -> bool:
        return (self.stock_nuevo is not None) or (self.precio_venta_nuevo is not None)


class ModeloActualizarProductos(QAbstractTableModel):
    def __init__(self, registros: List[RegistroProducto]):
        super().__init__()
        self._registros = registros

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._registros)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(HEADERS_UI)

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):
        if not index.isValid():
            return None

        r = self._registros[index.row()]
        fila = r.como_fila()
        valor = fila[index.column()]

        if role in (Qt.DisplayRole, Qt.EditRole):
            return valor

        if role == Qt.TextAlignmentRole:
            if index.column() in (
                COL_STOCK_ACTUAL, COL_STOCK_NUEVO,
                COL_PRECIO_ACTUAL, COL_PRECIO_COMPRA,
                COL_PRECIO_VENTA_ACTUAL, COL_PRECIO_VENTA_NUEVO,
                COL_ESTADO
            ):
                return Qt.AlignCenter
            return Qt.AlignVCenter

        return None

    def flags(self, index: QModelIndex):
        flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if index.column() in (COL_STOCK_NUEVO, COL_PRECIO_VENTA_NUEVO):
            flags |= Qt.ItemIsEditable
        return flags

    def setData(self, index: QModelIndex, value, role: int = Qt.EditRole) -> bool:
        if not index.isValid() or role != Qt.EditRole:
            return False

        r = self._registros[index.row()]

        try:
            if index.column() == COL_STOCK_NUEVO:
                txt = str(value).strip()
                r.stock_nuevo = None if txt == "" else int(float(txt))
                if r.stock_nuevo is not None and r.stock_nuevo < 0:
                    return False

            elif index.column() == COL_PRECIO_VENTA_NUEVO:
                txt = str(value).strip().replace(",", ".")
                r.precio_venta_nuevo = None if txt == "" else float(txt)
                if r.precio_venta_nuevo is not None and r.precio_venta_nuevo < 0:
                    return False
            else:
                return False
        except Exception:
            return False

        # actualizar estado local
        r.estado = "Pendiente" if r.tiene_cambios() else "Sin cambio"

        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        return True

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if orientation == Qt.Horizontal:
            if role == Qt.DisplayRole:
                return HEADERS_UI[section]
            if role == Qt.FontRole:
                f = QFont()
                f.setBold(True)
                return f
            if role == Qt.TextAlignmentRole:
                return Qt.AlignCenter
        return None

    def actualizar_estado(self, row: int, estado: str):
        if 0 <= row < len(self._registros):
            self._registros[row].estado = estado
            idx = self.index(row, COL_ESTADO)
            self.dataChanged.emit(idx, idx, [Qt.DisplayRole])


class ControladorActualizarProductos:
    def __init__(self):
        self.cliente = ClienteWooCommerce()
        self.simples: List[RegistroProducto] = []
        self.variados: List[RegistroProducto] = []
        self.modelo_simples: Optional[ModeloActualizarProductos] = None
        self.modelo_variados: Optional[ModeloActualizarProductos] = None

    # -------- CARGAR ARCHIVO --------
    def cargar_archivo(self, ruta: str) -> Dict[str, Dict[str, Optional[float]]]:
        """
        Devuelve:
        sku -> {'STOCK': Optional[int], 'PRECIO': Optional[float]}
        - Si viene vacío: None (sin cambio)
        - Acepta alias de encabezados
        - En Excel valida encabezados requeridos en NEGRITA.
        """
        data: Dict[str, Dict[str, Optional[float]]] = {}

        ruta_lower = ruta.lower()
        if ruta_lower.endswith(".xls"):
            raise ValueError("Archivos .xls no son soportados. Guarda como .xlsx y vuelve a intentar.")

        SKU_ALIASES = ["SKU"]
        STOCK_ALIASES = ["STOCK", "STOCK NUEVO", "NUEVO STOCK"]
        PRECIO_ALIASES = ["PRECIO", "PRECIO VENTA", "PRECIO_VENTA", "PRECIO VENTA NUEVO", "PRECIO_VENTA_NUEVO"]

        if ruta_lower.endswith(".csv"):
            with open(ruta, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    raise ValueError("El CSV no tiene encabezados.")
                headers = [_norm_header(h) for h in reader.fieldnames]

                idx_sku_name = _mapear_columna(headers, SKU_ALIASES, True, "SKU")
                idx_stock_name = _mapear_columna(headers, STOCK_ALIASES, False, "STOCK")
                idx_precio_name = _mapear_columna(headers, PRECIO_ALIASES, True, "PRECIO")

                name_sku = reader.fieldnames[idx_sku_name]
                name_stock = reader.fieldnames[idx_stock_name] if idx_stock_name is not None else None
                name_precio = reader.fieldnames[idx_precio_name]

                for r in reader:
                    sku = (r.get(name_sku) or "").strip()
                    if not sku:
                        continue

                    stock = _leer_num_opcional(r.get(name_stock)) if name_stock else None
                    precio = _leer_num_opcional(r.get(name_precio))

                    data[sku] = {
                        "STOCK": None if stock is None else int(stock),
                        "PRECIO": None if precio is None else float(precio),
                    }

        else:
            wb = load_workbook(ruta, data_only=True)
            ws = wb.active

            raw_headers = [c.value for c in ws[1]]
            headers = [_norm_header(h) for h in raw_headers]

            idx_sku = _mapear_columna(headers, SKU_ALIASES, True, "SKU")
            idx_stock = _mapear_columna(headers, STOCK_ALIASES, False, "STOCK")
            idx_precio = _mapear_columna(headers, PRECIO_ALIASES, True, "PRECIO")

            _validar_negrita_excel(
                ws,
                idxs_requeridos=[idx_sku, idx_precio],
                nombres_requeridos=["SKU", "PRECIO"],
            )

            for row in ws.iter_rows(min_row=2, values_only=True):
                sku = str(row[idx_sku] or "").strip()
                if not sku:
                    continue

                stock = _leer_num_opcional(row[idx_stock]) if idx_stock is not None else None
                precio = _leer_num_opcional(row[idx_precio])

                data[sku] = {
                    "STOCK": None if stock is None else int(stock),
                    "PRECIO": None if precio is None else float(precio),
                }

        return data

    # -------- PROCESAR --------
    def procesar_productos(
        self,
        datos_archivo: Dict[str, Dict[str, Optional[float]]],
        callback: Optional[Callable[[int, str], None]] = None
    ) -> Tuple[ModeloActualizarProductos, ModeloActualizarProductos]:
        productos = self.cliente.obtener_productos()

        self.simples.clear()
        self.variados.clear()

        total = max(len(productos), 1)

        for i, p in enumerate(productos, start=1):
            sku = (p.get("sku") or "").strip()
            excel = datos_archivo.get(sku) if sku else None

            stock_nuevo = excel.get("STOCK") if excel else None
            precio_nuevo = excel.get("PRECIO") if excel else None

            tipo = str(p.get("type") or "simple").strip().lower()
            categoria_parent = ", ".join(c.get("name", "") for c in (p.get("categories") or []))

            if tipo == "simple":
                costo = _get_purchase_cost(p)
                reg = RegistroProducto(
                    sku=sku,
                    nombre=p.get("name", ""),
                    categoria=categoria_parent,
                    stock_actual=p.get("stock_quantity", 0),
                    stock_nuevo=stock_nuevo,
                    precio_actual=_fmt_money(p.get("price", 0)),
                    precio_compra=_fmt_money(costo),
                    precio_venta_actual=_fmt_money(p.get("regular_price", p.get("price", 0))),
                    precio_venta_nuevo=None if precio_nuevo is None else float(_fmt_money(precio_nuevo)),
                    estado="Pendiente" if (stock_nuevo is not None or precio_nuevo is not None) else "Sin cambio",
                    _id=p.get("id"),
                    _parent_id=None,
                    _tipo="simple",
                )
                self.simples.append(reg)

            elif tipo == "variable":
                parent_id = p.get("id")
                parent_name = p.get("name", "")
                variaciones = self.cliente.obtener_variaciones_producto(int(parent_id)) if parent_id else []
                for v in variaciones:
                    sku_v = (v.get("sku") or "").strip()
                    excel_v = datos_archivo.get(sku_v) if sku_v else None

                    stock_v = excel_v.get("STOCK") if excel_v else None
                    precio_v = excel_v.get("PRECIO") if excel_v else None

                    attrs = []
                    for a in (v.get("attributes") or []):
                        n = a.get("name") or ""
                        o = a.get("option") or ""
                        if n and o:
                            attrs.append(f"{n}: {o}")
                    variacion_txt = " | ".join(attrs)
                    nombre_mostrar = parent_name if not variacion_txt else f"{parent_name} ({variacion_txt})"

                    costo_v = _get_purchase_cost(v)

                    regv = RegistroProducto(
                        sku=sku_v,
                        nombre=nombre_mostrar,
                        categoria=categoria_parent,
                        stock_actual=v.get("stock_quantity", 0),
                        stock_nuevo=stock_v,
                        precio_actual=_fmt_money(v.get("price", 0)),
                        precio_compra=_fmt_money(costo_v),
                        precio_venta_actual=_fmt_money(v.get("regular_price", v.get("price", 0))),
                        precio_venta_nuevo=None if precio_v is None else float(_fmt_money(precio_v)),
                        estado="Pendiente" if (stock_v is not None or precio_v is not None) else "Sin cambio",
                        _id=v.get("id"),
                        _parent_id=parent_id,
                        _tipo="variation",
                    )
                    self.variados.append(regv)

            else:
                costo = _get_purchase_cost(p)
                reg = RegistroProducto(
                    sku=sku,
                    nombre=p.get("name", ""),
                    categoria=categoria_parent,
                    stock_actual=p.get("stock_quantity", 0),
                    stock_nuevo=stock_nuevo,
                    precio_actual=_fmt_money(p.get("price", 0)),
                    precio_compra=_fmt_money(costo),
                    precio_venta_actual=_fmt_money(p.get("regular_price", p.get("price", 0))),
                    precio_venta_nuevo=None if precio_nuevo is None else float(_fmt_money(precio_nuevo)),
                    estado=f"⚠ Tipo: {tipo}",
                    _id=p.get("id"),
                    _parent_id=None,
                    _tipo=tipo,
                )
                self.variados.append(reg)

            if callback:
                callback(int((i / total) * 100), f"Procesando producto {i} de {total}")

        self.modelo_simples = ModeloActualizarProductos(self.simples)
        self.modelo_variados = ModeloActualizarProductos(self.variados)
        return self.modelo_simples, self.modelo_variados

    # -------- APLICAR --------
    def aplicar_cambios(self, callback: Optional[Callable[[int, str], None]] = None):
        trabajos: List[Tuple[ModeloActualizarProductos, int, RegistroProducto]] = []

        if self.modelo_simples:
            for idx, r in enumerate(self.simples):
                if r.tiene_cambios():
                    trabajos.append((self.modelo_simples, idx, r))

        if self.modelo_variados:
            for idx, r in enumerate(self.variados):
                if r.tiene_cambios():
                    trabajos.append((self.modelo_variados, idx, r))

        total = max(len(trabajos), 1)

        for i, (modelo, row, r) in enumerate(trabajos, start=1):
            if not r._id:
                modelo.actualizar_estado(row, "❌ Sin ID")
                continue

            if r._tipo == "variation":
                if not r._parent_id:
                    modelo.actualizar_estado(row, "❌ Sin ID padre")
                else:
                    try:
                        self.cliente.actualizar_variacion(
                            producto_id=int(r._parent_id),
                            variacion_id=int(r._id),
                            stock=r.stock_nuevo if r.stock_nuevo is not None else None,
                            precio=r.precio_venta_nuevo if r.precio_venta_nuevo is not None else None,
                        )
                        modelo.actualizar_estado(row, "OK Actualizado")
                    except Exception as e:
                        modelo.actualizar_estado(row, f"❌ {str(e)[:60]}")
                if callback:
                    callback(int((i / total) * 100), f"Aplicando cambios {i} de {total}")
                continue

            if r._tipo != "simple":
                modelo.actualizar_estado(row, f"⚠ No editable ({r._tipo})")
                if callback:
                    callback(int((i / total) * 100), f"Aplicando cambios {i} de {total}")
                continue

            try:
                self.cliente.actualizar_producto(
                    producto_id=int(r._id),
                    stock=r.stock_nuevo if r.stock_nuevo is not None else None,
                    precio=r.precio_venta_nuevo if r.precio_venta_nuevo is not None else None,
                )
                modelo.actualizar_estado(row, "OK Actualizado")
            except Exception as e:
                modelo.actualizar_estado(row, f"❌ {str(e)[:60]}")

            if callback:
                callback(int((i / total) * 100), f"Aplicando cambios {i} de {total}")

    # -------- EXPORTAR --------
    def exportar_excel(self, ruta: str):
        """
        Exporta en DOS pestañas (como reporte):
        - Productos Simples
        - Productos Variados

        Columnas exportadas:
        SKU | NOMBRE | CATEGORÍA | STOCK (NUEVO) | PRECIO COMPRA | PRECIO VENTA (NUEVO) | ESTADO
        """
        HEADERS_EXPORT = [
            "SKU",
            "NOMBRE",
            "CATEGORÍA",
            "STOCK",
            "PRECIO COMPRA",
            "PRECIO VENTA",
            "ESTADO",
        ]

        def fila_export(r: RegistroProducto):
            return [
                r.sku,
                r.nombre,
                r.categoria,
                "" if r.stock_nuevo is None else int(r.stock_nuevo),
                _to_float_any(r.precio_compra) if _to_float_any(r.precio_compra) is not None else 0.0,
                "" if r.precio_venta_nuevo is None else float(r.precio_venta_nuevo),
                r.estado,
            ]

        def aplicar_formato(ws):
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="D9E1F2")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.append(HEADERS_EXPORT)
            for col in range(1, len(HEADERS_EXPORT) + 1):
                c = ws.cell(row=1, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center

            ws.freeze_panes = "A2"
            last_col = get_column_letter(len(HEADERS_EXPORT))
            ws.auto_filter.ref = f"A1:{last_col}1"

            widths = [18, 48, 30, 10, 14, 14, 18]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # STOCK entero
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = "0"
                    cell.alignment = center

            # precios 2 decimales
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
                for cell in row:
                    cell.number_format = "0.00"
                    cell.alignment = center

            # estado centrado
            for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
                for cell in row:
                    cell.alignment = center

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Productos Simples"
        aplicar_formato(ws1)
        for r in self.simples:
            ws1.append(fila_export(r))

        ws2 = wb.create_sheet("Productos Variados")
        aplicar_formato(ws2)
        for r in self.variados:
            ws2.append(fila_export(r))

        wb.save(ruta)
